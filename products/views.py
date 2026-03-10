from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import load_workbook

from .forms import ApprovedFilterForm, ProductUploadForm
from .models import Product


@dataclass
class UploadResult:
    created_count: int
    updated_count: int
    updated_ids: list[int]
    errors: list[str]


def _parse_decimal(value) -> Decimal:
    if value is None or str(value).strip() == '':
        raise InvalidOperation('Missing price')
    parsed = Decimal(str(value))
    if parsed <= 0:
        raise InvalidOperation('Price must be positive')
    return parsed


def _parse_int(value, label: str) -> int:
    if value is None or str(value).strip() == '':
        raise ValueError(f'Missing {label}')
    parsed = int(str(value))
    if label == 'product_id' and parsed <= 0:
        raise ValueError('product_id must be positive')
    if label == 'quantity' and parsed < 0:
        raise ValueError('quantity cannot be negative')
    return parsed


def _process_excel(upload) -> UploadResult:
    workbook = load_workbook(upload, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return UploadResult(0, 0, [], ['Excel file is empty.'])

    header = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
    required = ['product_id', 'name', 'category', 'price', 'quantity']
    missing = [field for field in required if field not in header]
    if missing:
        return UploadResult(
            0,
            0,
            [],
            [f'Missing required columns: {", ".join(missing)}'],
        )

    index = {field: header.index(field) for field in required}

    created_count = 0
    updated_count = 0
    updated_ids: list[int] = []
    errors: list[str] = []

    with transaction.atomic():
        for row_num, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            try:
                product_id = _parse_int(row[index['product_id']], 'product_id')
                name = str(row[index['name']]).strip()
                category = str(row[index['category']]).strip()
                price = _parse_decimal(row[index['price']])
                quantity = _parse_int(row[index['quantity']], 'quantity')
                if not name or not category:
                    raise ValueError('Name and category are required')
            except (ValueError, InvalidOperation) as exc:
                errors.append(f'Row {row_num}: {exc}')
                continue

            product, created = Product.objects.get_or_create(
                product_id=product_id,
                defaults={
                    'name': name,
                    'category': category,
                    'price': price,
                    'quantity': quantity,
                    'status': Product.Status.DRAFT,
                },
            )
            if created:
                created_count += 1
            else:
                product.name = name
                product.category = category
                product.price = price
                product.quantity = quantity
                product.status = Product.Status.DRAFT
                product.save()
                updated_count += 1
            updated_ids.append(product_id)

    return UploadResult(created_count, updated_count, updated_ids, errors)


def draft_upload(request):
    updated_ids = request.session.pop('updated_ids', [])
    upload_form = ProductUploadForm()

    if request.method == 'POST':
        upload_form = ProductUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            result = _process_excel(upload_form.cleaned_data['file'])

            if result.errors:
                preview_errors = '\n'.join(result.errors[:5])
                messages.error(
                    request,
                    'Some rows were skipped:\n' + preview_errors,
                )

            messages.success(
                request,
                f'Upload complete. Created: {result.created_count}, Updated: {result.updated_count}.',
            )
            request.session['updated_ids'] = result.updated_ids
            return redirect('products:draft_upload')
        messages.error(request, 'Please upload a valid .xlsx file.')

    drafts = Product.objects.filter(status=Product.Status.DRAFT)
    paginator = Paginator(drafts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'upload_form': upload_form,
        'page_obj': page_obj,
        'updated_ids': updated_ids,
    }
    return render(request, 'products/draft_upload.html', context)


@require_POST
def approve_product(request, product_id: int):
    product = get_object_or_404(Product, product_id=product_id)
    product.status = Product.Status.APPROVED
    product.save()
    messages.success(request, f'Product {product.product_id} approved.')
    return redirect('products:draft_upload')


def approved_products(request):
    form = ApprovedFilterForm(request.GET or None)
    products = Product.objects.filter(status=Product.Status.APPROVED)

    per_page = 10
    if form.is_valid():
        query = form.cleaned_data.get('q')
        if query:
            products = products.filter(Q(name__icontains=query) | Q(category__icontains=query))

        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        if start_date:
            products = products.filter(last_updated__date__gte=start_date)
        if end_date:
            products = products.filter(last_updated__date__lte=end_date)

        per_page_value = form.cleaned_data.get('per_page')
        if per_page_value:
            per_page = int(per_page_value)

    paginator = Paginator(products, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    query_string = query_params.urlencode()

    context = {
        'form': form,
        'page_obj': page_obj,
        'query_string': query_string,
    }
    return render(request, 'products/approved_list.html', context)
